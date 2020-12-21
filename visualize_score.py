import json
import os
import string

import numpy as np
from openpyxl import Workbook

upper = string.ascii_uppercase
alphabet = (
    [el for el in upper] + ["A" + el for el in upper] + ["B" + el for el in upper]
)

fname = "./score/dd/valid.jsonl"
threshold = 0.3

with open(fname, "r") as f:
    ls = [json.loads(el) for el in f.readlines()][:100]

wb = Workbook()
ws = wb.active

print(ls[0].keys())

for idx, item in enumerate(ls):
    context, reply, tokenized_reply, response_score, conv_score = (
        item["context"],
        item["reply"],
        item["tokenized_reply"],
        item["response_score"],
        item["conv_score"],
    )
    begin_row = 1 + idx * 10

    ws["A" + str(1 + idx * 10)] = "Context"
    ws["B" + str(1 + idx * 10)] = " ||| ".join(context)
    ws["A" + str(2 + idx * 10)] = "Reply"
    ws["B" + str(2 + idx * 10)] = reply
    ws["A" + str(4 + idx * 10)] = "Conv"
    ws["A" + str(5 + idx * 10)] = "Res"
    ws["A" + str(6 + idx * 10)] = "Minus"
    ws["A" + str(7 + idx * 10)] = "Con"
    ws["A" + str(8 + idx * 10)] = "Res"
    ws["A" + str(9 + idx * 10)] = "Minus"

    diff_score_list = []
    for tok_id, tok in enumerate(tokenized_reply):
        diff_score = round(conv_score[tok_id] - response_score[tok_id], 4)
        diff_score_list.append(diff_score)
        ws[alphabet[tok_id + 1] + str(3 + idx * 10)] = tok
        ws[alphabet[tok_id + 1] + str(4 + idx * 10)] = round(conv_score[tok_id], 4)
        ws[alphabet[tok_id + 1] + str(5 + idx * 10)] = round(response_score[tok_id], 4)
        ws[alphabet[tok_id + 1] + str(6 + idx * 10)] = diff_score
        ws[alphabet[tok_id + 1] + str(7 + idx * 10)] = ""
        ws[alphabet[tok_id + 1] + str(8 + idx * 10)] = ""

    max_score = np.argsort(np.array(diff_score_list))[-1]
    for tok_id, tok in enumerate(tokenized_reply):
        ws[alphabet[tok_id + 1] + str(9 + idx * 10)] = (
            "[MASK]" if max_score == tok_id else tok
        )


os.makedirs("image", exist_ok=True)
wb.save("image/valid.xlsx")
