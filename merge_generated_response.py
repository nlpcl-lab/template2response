import os, json, pickle
import argparse

parser = argparse.ArgumentParser(
    description="Configuration for template generation"
)
parser.add_argument(
    "--dataset", type=str, default="dd", choices=["dd", "persona"]
)

args = parser.parse_args()

MODEL_DICT = {
    "Infill_greedy": "./logs/gpt2_infill-{}/greedy_pppl100.jsonl".format(
        args.dataset
    ),
    "GPT2": "./model_generated/gpt2_{}_beam5.jsonl".format(args.dataset),
    "GPT2_refine": "./model_generated/gpt2_refine_{}_beam5.jsonl".format(
        args.dataset
    ),
    "Infill_beam": "./logs/gpt2_infill-{}/beam5_pppl100.jsonl".format(
        args.dataset
    ),
    "Infill_random_beam": "./logs/gpt2_infill-{}/beam_random.jsonl".format(
        args.dataset
    ),
}


assert all([os.path.exists(val) for key, val in MODEL_DICT.items()])
data_dict = {}
for model_name, fname in MODEL_DICT.items():

    with open(fname, "r") as f:
        data = [json.loads(el) for el in f.readlines()][4:]

    for idx, line in enumerate(data):
        if "context" not in line:
            continue
        if "final_output" not in line:
            continue
        if "Infill" in model_name:
            retrieved = line["retrieved"]

        context, reply, generated = (
            line["context"],
            line["golden"],
            line["final_output"],
        )
        if "refine" in model_name:
            context = context.split("[RTV]")[1]
        if isinstance(context, str):
            context = [
                el.strip()
                for el in context.split("[SEPT]")
                if len(el.strip()) != 0
            ]
        generated = generated.replace("<|endoftext|>", "").strip()
        generated = generated.replace(" ##", "")

        if args.dataset == "dd":
            if " ".join(context) not in data_dict:
                data_dict[" ".join(context)] = {}
                data_dict[" ".join(context)]["context"] = context
                data_dict[" ".join(context)]["id"] = idx
                data_dict[" ".join(context)]["golden"] = reply
                data_dict[" ".join(context)]["generated"] = {
                    model_name: generated
                }
                if "Infill" in model_name:
                    data_dict[" ".join(context)]["retrieved"] = retrieved

            else:
                if model_name in data_dict[" ".join(context)]["generated"]:
                    continue
                data_dict[" ".join(context)]["generated"][
                    model_name
                ] = generated
        else:
            if idx not in data_dict:
                data_dict[idx] = {}
                data_dict[idx]["context"] = context
                data_dict[idx]["id"] = idx
                data_dict[idx]["golden"] = reply
                data_dict[idx]["generated"] = {model_name: generated}
                if "Infill" in model_name:
                    data_dict[idx]["retrieved"] = retrieved
            else:
                assert (
                    context[-1] in data_dict[idx]["context"][-1]
                    or data_dict[idx]["context"][-1] in context[-1]
                )

                data_dict[idx]["generated"][model_name] = generated


del_key_list = []
for idx, item in data_dict.items():
    if "generated" not in item:
        del_key_list.append(idx)
    elif len(item["generated"]) != len(list(MODEL_DICT.keys())):
        del_key_list.append(idx)
for key in del_key_list:
    del data_dict[key]

merged_data = []
for _, item in data_dict.items():
    context, golden, generated = (
        item["context"],
        item["golden"],
        item["generated"],
    )
    merged_data.append(
        {
            "context": context,
            "golden": golden,
            "model": generated,
            "retrieved": item["retrieved"],
        }
    )
del data_dict
print(len(merged_data))

with open("./generated/{}_result.jsonl".format(args.dataset), "w") as f:
    for line in merged_data:
        json.dump(line, f)
        f.write("\n")

from openpyxl import Workbook
from openpyxl.styles import Font

wb = Workbook()
ws = wb.active

counter = 1
for idx, item in enumerate(merged_data):
    context, golden, generated = (
        item["context"],
        item["golden"],
        item["model"],
    )
    retrieved = item["retrieved"]
    ws["A" + str(counter)] = idx
    ws["A" + str(counter)].font = Font(bold=True)
    counter += 1

    ws["A" + str(counter)] = "Context"
    ws["A" + str(counter)].font = Font(bold=True)
    for context_idx, turn in enumerate(context):
        ws["B" + str(counter)] = turn
        counter += 1
    ws["A" + str(counter)] = "Golden"
    ws["A" + str(counter)].font = Font(bold=True)
    ws["B" + str(counter)] = golden
    counter += 2
    ws["A" + str(counter)] = "Retrieved"
    ws["A" + str(counter)].font = Font(bold=True)
    ws["B" + str(counter)] = retrieved
    counter += 1
    for model in generated:
        ws["A" + str(counter)] = model
        ws["A" + str(counter)].font = Font(bold=True)
        ws["B" + str(counter)] = generated[model]
        counter += 1
    counter += 1
wb.save("./generated/{}_result.xlsx".format(args.dataset))
