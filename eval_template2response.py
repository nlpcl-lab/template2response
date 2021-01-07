import json
import os
import string

import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font

upper = string.ascii_uppercase
alphabet = (
    [el for el in upper]
    + ["A" + el for el in upper]
    + ["B" + el for el in upper]
    + ["C" + el for el in upper]
    + ["D" + el for el in upper]
    + ["E" + el for el in upper]
    + ["F" + el for el in upper]
)


dataset = "dd"
threshold = 0.5
setname = "test"

fname = "./generated/{}-thres{}-decode_argmax_onestep_{}.jsonl".format(
    dataset, threshold, setname
)


assert os.path.exists(fname)
with open(fname, "r") as f:
    ls = [json.loads(el) for el in f.readlines()]

wb = Workbook()
ws = wb.active

print(ls[0].keys())

row_counter = 1


for idx, item in enumerate(ls):
    if item == {}:
        continue
    print(f"{idx}/{len(ls)}")
    (
        context,
        retrieved,
        golden,
        context_score,
        response_score,
        context_changed,
        response_changed,
    ) = (
        item["context"],
        item["retrieved"],
        item["golden"],
        item["context_score"],
        item["response_score"],
        item["context_output"],
        item["response_output"],
    )

    assert (
        len(context_changed[0].split())
        == len(response_score)
        == len(context_score)
        == len(response_changed[0].split())
    )
    ws["A" + str(row_counter)] = "Context"
    ws["A" + str(row_counter)].font = Font(bold=True)
    speaker = "A"
    for uttr in context:
        ws["B" + str(row_counter)] = speaker + ": " + uttr
        speaker = "B" if speaker == "A" else "A"
        row_counter += 1
    ws["A" + str(row_counter)] = "Golden"
    ws["A" + str(row_counter)].font = Font(bold=True)
    ws["B" + str(row_counter)] = golden
    row_counter += 1
    ws["A" + str(row_counter)] = "Retrieved"
    ws["A" + str(row_counter)].font = Font(bold=True)
    ws["B" + str(row_counter)] = retrieved
    row_counter += 1

    """
    Related with Response Score
    """
    ws["A" + str(row_counter)] = "Response-Skeleton"
    ws["A" + str(row_counter)].font = Font(bold=True)
    ws["B" + str(row_counter)] = response_changed[0]
    row_counter += 1

    ws["A" + str(row_counter)] = "Score"
    try:
        for tok_id, tok in enumerate(retrieved.split()):
            ws[alphabet[tok_id + 1] + str(row_counter)] = tok
            ws[alphabet[tok_id + 1] + str(row_counter + 1)] = round(
                response_score[tok_id], 3
            )
    except:
        continue
    row_counter += 3

    response_stage = [
        tok if response_score[tok_id] > threshold else "__"
        for tok_id, tok in enumerate(response_changed[0].split())
    ]
    ws["A" + str(row_counter)] = "STEP0"
    for tok_id, tok in enumerate(response_changed[0].split()):
        ws[alphabet[tok_id + 1] + str(row_counter)] = response_stage[tok_id]
    row_counter += 1

    for step, (index, word) in enumerate(response_changed[1]):
        response_stage[index] = word
        ws["A" + str(row_counter)] = "STEP{}".format(step + 1)
        for tok_id, tok in enumerate(response_stage):
            ws[alphabet[tok_id + 1] + str(row_counter)] = tok
            if tok_id == index:
                ws[alphabet[tok_id + 1] + str(row_counter)].font = Font(
                    bold=True
                )
        row_counter += 1
    row_counter += 1

    """
    Related with Context Score
    """
    ws["A" + str(row_counter)] = "Context-Skeleton"
    ws["A" + str(row_counter)].font = Font(bold=True)
    ws["B" + str(row_counter)] = context_changed[0]
    row_counter += 1

    ws["A" + str(row_counter)] = "Score"
    for tok_id, tok in enumerate(retrieved.split()):
        ws[alphabet[tok_id + 1] + str(row_counter)] = tok
        ws[alphabet[tok_id + 1] + str(row_counter + 1)] = round(
            context_score[tok_id], 3
        )
    row_counter += 3

    response_stage = [
        tok if context_score[tok_id] > threshold else "__"
        for tok_id, tok in enumerate(context_changed[0].split())
    ]
    ws["A" + str(row_counter)] = "STEP0"
    for tok_id, tok in enumerate(context_changed[0].split()):
        ws[alphabet[tok_id + 1] + str(row_counter)] = response_stage[tok_id]
    row_counter += 1

    for step, (index, word) in enumerate(context_changed[1]):
        response_stage[index] = word
        ws["A" + str(row_counter)] = "STEP{}".format(step + 1)
        for tok_id, tok in enumerate(response_stage):
            ws[alphabet[tok_id + 1] + str(row_counter)] = tok
            if tok_id == index:
                ws[alphabet[tok_id + 1] + str(row_counter)].font = Font(
                    bold=True
                )
        row_counter += 1
    row_counter += 1


wb.save(fname.replace(".jsonl", ".xlsx"))
